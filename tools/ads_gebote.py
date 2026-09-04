#!/usr/bin/env python3
"""
ads_gebote.py — Gebote fuer Sponsored Products lesen und setzen (nur Coach, nur lokal).

Ruft die Edge Function `ads-gebote` mit DEINER Supabase-Session auf. Du bist damit
derselbe Nutzer wie in der Coach-Ansicht des Portals; die Function laesst nur
Plattform-Admins durch. ChatGPT/MCP haben diesen Weg nicht.

Ablauf:
  1. Einmalig anmelden (Passwort wird abgefragt, nie gespeichert; nur die Session
     landet in tools/.op_session.json, die Datei ist in .gitignore):
        python tools/ads_gebote.py login

  2. Kampagnen ansehen:
        python tools/ads_gebote.py kampagnen --firma Vaneja
        python tools/ads_gebote.py kampagnen --firma Vaneja --status ENABLED,PAUSED,ARCHIVED

  3. Gebote einer oder mehrerer Kampagnen ansehen (ID oder Namensteil, mehrfach):
        python tools/ads_gebote.py gebote --firma Vaneja --kampagne "Brand" --kampagne 12345

  4. Vorschau einer Aenderung (schreibt NICHTS, legt tools/.vorschau.json ab):
        python tools/ads_gebote.py vorschau --firma Vaneja --kampagne "Brand" --prozent -20
        python tools/ads_gebote.py vorschau --firma Vaneja --kampagne "Brand" --faktor 0.8 --min 0.25
        python tools/ads_gebote.py vorschau --firma Vaneja --kampagne "Brand" --absolut 0.45 --nur keyword

  5. Genau diese Vorschau anwenden (fragt noch einmal nach, ausser mit --ja):
        python tools/ads_gebote.py setzen --firma Vaneja --grund "ACOS zu hoch"

  Alternativ eine Excel-Tabelle einlesen (Blatt mit Spalten Entity, Campaign ID,
  Keyword-ID / Targeting-ID, Keyword / Target, Zustand, Gebot alt, Gebot NEU,
  Begruendung). Zeilen mit Zustand "paused"/"enabled" aendern auch den Zustand.
  Ergebnis ist dieselbe Vorschau wie oben, danach normal "setzen":
        python tools/ads_gebote.py tabelle --firma Vaneja --datei "Massnahmen.xlsx" --blatt Gebotsänderungen
     Ohne --datei wird die neueste .xlsx aus tools/eingang/ genommen.

Wo ausfuehren: auf dem PC (nicht VPS). Braucht nur Python 3 + requests.
"""

import argparse
import getpass
import json
import os
import sys
import time

import requests

# Windows-Konsole: Umlaute sauber ausgeben
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SUPABASE_URL = "https://irvnghhxfjjnpodclfuc.supabase.co"
# oeffentlicher anon-Key (kein Geheimnis, siehe FRONTEND.md)
ANON_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imlydm5naGh4ZmpqbnBvZGNsZnVjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODQyMzUzNDAsImV4cCI6MjA5OTgxMTM0MH0."
    "lUSreFAj7oB5dN_KO2GKrb5s6COaUCiQm0_GZuQYJtM"
)
HIER = os.path.dirname(os.path.abspath(__file__))
SESSION_DATEI = os.path.join(HIER, ".op_session.json")
VORSCHAU_DATEI = os.path.join(HIER, ".vorschau.json")
FUNKTION = f"{SUPABASE_URL}/functions/v1/ads-gebote"


# ----------------------------------------------------------------- Session

def _speichere_session(d):
    with open(SESSION_DATEI, "w", encoding="utf-8") as f:
        json.dump({
            "access_token": d["access_token"],
            "refresh_token": d["refresh_token"],
            "expires_at": int(time.time()) + int(d.get("expires_in", 3600)),
            "email": (d.get("user") or {}).get("email"),
        }, f)


def login(args):
    email = args.email or input("E-Mail [info@tl-ecom.de]: ").strip() or "info@tl-ecom.de"
    pw = getpass.getpass("Passwort: ")
    r = requests.post(
        f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
        headers={"apikey": ANON_KEY, "Content-Type": "application/json"},
        json={"email": email, "password": pw}, timeout=30,
    )
    if r.status_code != 200:
        sys.exit(f"Anmeldung fehlgeschlagen ({r.status_code}): {r.text[:300]}")
    _speichere_session(r.json())
    print(f"Angemeldet als {email}. Session liegt in {SESSION_DATEI}")


def token():
    if not os.path.exists(SESSION_DATEI):
        sys.exit("Keine Session. Erst:  python tools/ads_gebote.py login")
    with open(SESSION_DATEI, encoding="utf-8") as f:
        s = json.load(f)
    if s.get("expires_at", 0) - 60 > time.time():
        return s["access_token"]
    r = requests.post(
        f"{SUPABASE_URL}/auth/v1/token?grant_type=refresh_token",
        headers={"apikey": ANON_KEY, "Content-Type": "application/json"},
        json={"refresh_token": s["refresh_token"]}, timeout=30,
    )
    if r.status_code != 200:
        sys.exit(f"Session abgelaufen, bitte neu anmelden (login). {r.text[:200]}")
    d = r.json()
    _speichere_session(d)
    return d["access_token"]


def ruf(body):
    r = requests.post(
        FUNKTION,
        headers={"Authorization": f"Bearer {token()}", "apikey": ANON_KEY, "Content-Type": "application/json"},
        json=body, timeout=300,
    )
    try:
        d = r.json()
    except ValueError:
        sys.exit(f"Unerwartete Antwort ({r.status_code}): {r.text[:300]}")
    if r.status_code >= 400:
        sys.exit(f"Fehler {r.status_code}: {json.dumps(d, ensure_ascii=False)[:800]}")
    return d


# ----------------------------------------------------------------- Aufloesen

def firma_id(name):
    firmen = ruf({"action": "firmen"})["firmen"]
    treffer = [f for f in firmen if f["name"] and name.lower() in f["name"].lower()]
    if len(treffer) == 1:
        return treffer[0]["tenant_id"], treffer[0]["name"]
    if not treffer:
        sys.exit("Keine Firma mit Ads-Verbindung passt zu '%s'. Vorhanden: %s"
                 % (name, ", ".join(f["name"] or "?" for f in firmen)))
    sys.exit("Mehrdeutig: " + ", ".join(f["name"] for f in treffer))


def kampagnen_ids(tenant, auswahl, status):
    alle = ruf({"action": "kampagnen", "company_id": tenant, "status": status})["kampagnen"]
    ids, namen = [], []
    for a in auswahl:
        exakt = [k for k in alle if k["campaignId"] == a]
        treffer = exakt or [k for k in alle if a.lower() in (k["name"] or "").lower()]
        if not treffer:
            sys.exit(f"Keine Kampagne passt zu '{a}' (Status-Filter: {status}).")
        for k in treffer:
            if k["campaignId"] not in ids:
                ids.append(k["campaignId"])
                namen.append(k["name"])
    return ids, namen, {k["campaignId"]: k["name"] for k in alle}


# ----------------------------------------------------------------- Filter

def filtere(zeilen, treffer):
    """--treffer: nur Zeilen, deren id ODER text (Teilstring, ohne Gross/Klein) passt."""
    if not treffer:
        return zeilen
    out = []
    for z in zeilen:
        for t in treffer:
            if z["id"] == t or t.lower() in (z.get("text") or "").lower():
                out.append(z)
                break
    return out


# ----------------------------------------------------------------- Ausgabe

def tabelle(zeilen, spalten):
    if not zeilen:
        print("(keine Zeilen)")
        return
    breiten = [max(len(str(s)), *(len(str(z.get(s, ""))) for z in zeilen)) for s in spalten]
    breiten = [min(b, 48) for b in breiten]
    kopf = "  ".join(str(s).ljust(b) for s, b in zip(spalten, breiten))
    print(kopf)
    print("-" * len(kopf))
    for z in zeilen:
        print("  ".join(str(z.get(s, ""))[:b].ljust(b) for s, b in zip(spalten, breiten)))


# ----------------------------------------------------------------- Befehle

def cmd_firmen(args):
    tabelle(ruf({"action": "firmen"})["firmen"], ["name", "tenant_id", "profile_id", "marketplace_id", "status"])


def cmd_kampagnen(args):
    tenant, name = firma_id(args.firma)
    d = ruf({"action": "kampagnen", "company_id": tenant, "status": args.status.split(",")})
    print(f"Firma: {name}   Ads-Profil: {d['profile_id']}   Kampagnen: {len(d['kampagnen'])}")
    tabelle(d["kampagnen"], ["campaignId", "name", "state", "targetingType", "budget", "strategie"])


def cmd_gebote(args):
    tenant, name = firma_id(args.firma)
    ids, namen, kname = kampagnen_ids(tenant, args.kampagne, args.kampagnen_status.split(","))
    d = ruf({"action": "gebote", "company_id": tenant, "kampagnen": ids, "status": args.status.split(","), "nur": args.nur})
    d["zeilen"] = filtere(d["zeilen"], args.treffer)
    for z in d["zeilen"]:
        z["kampagne"] = kname.get(z["campaignId"], z["campaignId"])
    print(f"Firma: {name}   Kampagnen: {', '.join(namen)}")
    print(f"Zeilen mit eigenem Gebot: {d['anzahl']}   erben Standardgebot der Anzeigengruppe: {d['erben_standard']}"
          + (f"   nach Filter: {len(d['zeilen'])}" if args.treffer else ""))
    tabelle(sorted(d["zeilen"], key=lambda z: (z["kampagne"], -z["gebot"])),
            ["kampagne", "art", "text", "matchType", "state", "gebot", "id"])


def cmd_vorschau(args):
    tenant, name = firma_id(args.firma)
    ids, namen, kname = kampagnen_ids(tenant, args.kampagne, args.kampagnen_status.split(","))
    regel = {k: getattr(args, k) for k in ("prozent", "faktor", "absolut", "min", "max") if getattr(args, k) is not None}
    d = ruf({"action": "vorschau", "company_id": tenant, "kampagnen": ids, "status": args.status.split(","),
             "nur": args.nur, "regel": regel})
    d["aenderungen"] = filtere(d["aenderungen"], args.treffer)
    for a in d["aenderungen"]:
        a["kampagne"] = kname.get(a["campaignId"], a["campaignId"])
    z = d["zusammenfassung"]
    if args.treffer:
        ae = d["aenderungen"]
        z = {"anzahl": len(ae), "keywords": sum(a["art"] == "keyword" for a in ae), "targets": sum(a["art"] == "target" for a in ae),
             "summe_alt": round(sum(a["gebot"] for a in ae), 2), "summe_neu": round(sum(a["neu"] for a in ae), 2)}
    print(f"Firma: {name}   Kampagnen: {', '.join(namen)}   Regel: {regel}")
    print(f"Aenderungen: {z['anzahl']} (Keywords {z['keywords']}, Targets {z['targets']})   "
          f"Summe Gebote {z['summe_alt']} -> {z['summe_neu']}   ohne eigenes Gebot (unveraendert): {d['erben_standard']}")
    tabelle(sorted(d["aenderungen"], key=lambda a: (a["kampagne"], -a["gebot"])),
            ["kampagne", "art", "text", "matchType", "gebot", "neu", "delta", "id"])
    with open(VORSCHAU_DATEI, "w", encoding="utf-8") as f:
        json.dump({"firma": name, "tenant_id": tenant, "kampagnen": namen, "regel": regel,
                   "erstellt": time.strftime("%Y-%m-%d %H:%M:%S"), "aenderungen": d["aenderungen"]}, f, ensure_ascii=False, indent=1)
    print(f"\nNICHTS geschrieben. Vorschau liegt in {VORSCHAU_DATEI}.")
    print("Anwenden:  python tools/ads_gebote.py setzen --firma %s --grund \"...\"" % args.firma)


def _zahl(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace("€", "").replace(" ", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def _id(v):
    """Excel liefert IDs gern als Float (337058135144638.0) -> ganzzahliger String."""
    if v is None:
        return None
    if isinstance(v, float):
        return str(int(v))
    t = str(v).strip()
    return t[:-2] if t.endswith(".0") else t


def lies_tabelle(datei, blatt):
    """Liest ein Massnahmen-Blatt. Spalten werden ueber die Kopfzeile gefunden (Teilstring,
    ohne Gross/Klein); die Kopfzeile ist die erste Zeile mit 'Gebot' und 'NEU'."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl fehlt:  pip install openpyxl")
    wb = openpyxl.load_workbook(datei, data_only=True)
    if blatt not in wb.sheetnames:
        sys.exit(f"Blatt '{blatt}' nicht gefunden. Vorhanden: {', '.join(wb.sheetnames)}")
    ws = wb[blatt]
    zeilen = list(ws.iter_rows(values_only=True))
    kopf_idx = next((i for i, r in enumerate(zeilen)
                     if any(c and "gebot" in str(c).lower() and "neu" in str(c).lower() for c in r)), None)
    if kopf_idx is None:
        sys.exit("Keine Kopfzeile mit 'Gebot NEU' gefunden.")
    kopf = [str(c).strip().lower() if c is not None else "" for c in zeilen[kopf_idx]]

    def spalte(*teile):
        for i, k in enumerate(kopf):
            if all(t in k for t in teile):
                return i
        return None

    sp = {
        "entity": spalte("entity"), "campaign": spalte("campaign"),
        "id": spalte("keyword-id") if spalte("keyword-id") is not None else spalte("targeting-id"),
        "text": spalte("keyword / target") if spalte("keyword / target") is not None else spalte("keyword"),
        "match": spalte("match"), "zustand": spalte("zustand") if spalte("zustand") is not None else spalte("state"),
        "alt": spalte("gebot", "alt"), "neu": spalte("gebot", "neu"),
        "grund": spalte("begr") if spalte("begr") is not None else spalte("grund"),
    }
    fehlend = [k for k in ("entity", "campaign", "id", "neu") if sp[k] is None]
    if fehlend:
        sys.exit(f"Spalten nicht gefunden: {fehlend}. Kopfzeile: {kopf}")

    def wert(r, k):
        i = sp.get(k)
        return r[i] if i is not None and i < len(r) else None

    out = []
    for r in zeilen[kopf_idx + 1:]:
        ent = wert(r, "entity")
        if not ent or not _id(wert(r, "id")):
            continue
        art = "keyword" if "keyword" in str(ent).lower() else "target"
        zustand = str(wert(r, "zustand") or "").strip().lower()
        state = "PAUSED" if zustand.startswith("paus") else ("ENABLED" if zustand.startswith(("enab", "akt")) else None)
        out.append({
            "art": art, "id": _id(wert(r, "id")), "campaignId": _id(wert(r, "campaign")),
            "text": str(wert(r, "text") or ""), "matchType": str(wert(r, "match") or ""),
            "gebot": _zahl(wert(r, "alt")), "neu": _zahl(wert(r, "neu")), "state": state,
            "grund": (str(wert(r, "grund")).strip() if wert(r, "grund") else None),
        })
    return out


EINGANG = os.path.join(HIER, "eingang")


def neueste_datei():
    """Ohne --datei: die zuletzt geaenderte .xlsx aus tools/eingang."""
    import glob
    dateien = [f for f in glob.glob(os.path.join(EINGANG, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if not dateien:
        sys.exit(f"Keine .xlsx in {EINGANG}. Datei dort ablegen oder --datei angeben.")
    return max(dateien, key=os.path.getmtime)


def cmd_tabelle(args):
    tenant, name = firma_id(args.firma)
    if not args.datei:
        args.datei = neueste_datei()
        print(f"Datei: {args.datei}")
    zeilen = lies_tabelle(args.datei, args.blatt)
    if not zeilen:
        sys.exit("Keine Datenzeilen im Blatt.")
    kname = {k["campaignId"]: k["name"] for k in
             ruf({"action": "kampagnen", "company_id": tenant, "status": ["ENABLED", "PAUSED", "ARCHIVED"]})["kampagnen"]}
    ist = ruf({"action": "pruefen", "company_id": tenant,
               "aenderungen": [{"art": z["art"], "id": z["id"]} for z in zeilen]})["zeilen"]
    by_id = {f"{z['art']}:{z['id']}": z for z in ist}

    aend, schon, problem = [], [], []
    for z in zeilen:
        a = by_id.get(f"{z['art']}:{z['id']}")
        z["kampagne"] = kname.get(z["campaignId"], z["campaignId"])
        if not a:
            z["hinweis"] = "bei Amazon nicht gefunden"; problem.append(z); continue
        z["ist"] = a["gebot"]; z["state_alt"] = a.get("state"); z["text"] = z["text"] or a["text"]
        gebot_aendert = z["neu"] is not None and abs(a["gebot"] - z["neu"]) > 0.005
        state_aendert = z["state"] is not None and a.get("state") != z["state"]
        if not gebot_aendert and not state_aendert:
            z["hinweis"] = "schon so"; schon.append(z); continue
        if gebot_aendert and z["gebot"] is not None and abs(a["gebot"] - z["gebot"]) > 0.005:
            z["hinweis"] = f"Tabelle sagt alt {z['gebot']}, Amazon hat {a['gebot']}"; problem.append(z); continue
        z["was"] = " + ".join(x for x in [
            f"Gebot {a['gebot']} -> {z['neu']}" if gebot_aendert else "",
            f"Zustand {a.get('state')} -> {z['state']}" if state_aendert else ""] if x)
        z["delta"] = round(z["neu"] - a["gebot"], 2) if gebot_aendert else ""
        if not gebot_aendert:
            z["neu"] = None
        if not state_aendert:
            z["state"] = None
        aend.append(z)

    print(f"Firma: {name}   Datei: {os.path.basename(args.datei)} / {args.blatt}   Zeilen: {len(zeilen)}")
    print(f"Anzuwenden: {len(aend)}   schon so: {len(schon)}   Probleme (werden NICHT geschrieben): {len(problem)}\n")
    if aend:
        print("ANZUWENDEN")
        tabelle(aend, ["kampagne", "art", "text", "matchType", "was", "delta", "id"])
    if schon:
        print("\nSCHON SO (uebersprungen)")
        tabelle(schon, ["kampagne", "art", "text", "ist", "state_alt", "id"])
    if problem:
        print("\nPROBLEME (uebersprungen)")
        tabelle(problem, ["kampagne", "art", "text", "hinweis", "id"])
    if not aend:
        sys.exit("\nNichts anzuwenden.")
    with open(VORSCHAU_DATEI, "w", encoding="utf-8") as f:
        json.dump({"firma": name, "tenant_id": tenant, "kampagnen": sorted({a["kampagne"] for a in aend}),
                   "regel": {"tabelle": os.path.basename(args.datei), "blatt": args.blatt},
                   "erstellt": time.strftime("%Y-%m-%d %H:%M:%S"), "aenderungen": aend}, f, ensure_ascii=False, indent=1)
    print(f"\nNICHTS geschrieben. Vorschau liegt in {VORSCHAU_DATEI}.")
    print("Anwenden:  python tools/ads_gebote.py setzen --firma %s" % args.firma)


def cmd_setzen(args):
    if not os.path.exists(VORSCHAU_DATEI):
        sys.exit("Keine Vorschau vorhanden. Erst:  python tools/ads_gebote.py vorschau ...")
    with open(VORSCHAU_DATEI, encoding="utf-8") as f:
        v = json.load(f)
    tenant, name = firma_id(args.firma)
    if tenant != v["tenant_id"]:
        sys.exit(f"Die Vorschau gehoert zu '{v['firma']}', nicht zu '{name}'. Abbruch.")
    aend = v["aenderungen"]
    if not aend:
        sys.exit("Die Vorschau enthaelt keine Aenderungen.")
    print(f"Firma: {name}   Kampagnen: {', '.join(v['kampagnen'])}   Regel: {v['regel']}   Vorschau von {v['erstellt']}")
    print(f"{len(aend)} Gebote werden bei Amazon gesetzt.")
    if not args.ja:
        antwort = input("Wirklich schreiben? (ja/nein): ").strip().lower()
        if antwort != "ja":
            sys.exit("Abgebrochen. Nichts geschrieben.")
    d = ruf({"action": "setzen", "company_id": tenant, "bestaetigung": True, "grund": args.grund,
             "aenderungen": [{"art": a["art"], "id": a["id"], "gebot": a.get("ist", a.get("gebot")), "neu": a.get("neu"),
                              "state": a.get("state"), "state_alt": a.get("state_alt"), "grund": a.get("grund")} for a in aend]})
    print(f"Geschrieben: {d['geschrieben']}   Fehler: {d['fehler']}   Uebersprungen: {d['uebersprungen']}")
    ok = [e for e in d["ergebnisse"] if e["ergebnis"] == "ok"]
    if ok:
        tabelle(ok, ["art", "text", "alt", "neu", "state_alt", "state", "id"])
    problem = [e for e in d["ergebnisse"] if e["ergebnis"] != "ok"]
    if problem:
        print("\nNicht geschrieben:")
        tabelle(problem, ["art", "id", "text", "alt", "neu", "state", "ergebnis", "detail"])
    if d.get("log_fehler"):
        print(f"WARNUNG: Log konnte nicht geschrieben werden: {d['log_fehler']}")
    os.remove(VORSCHAU_DATEI)


# ----------------------------------------------------------------- Weitere Aktionen
# (Platzierungs-Modifier, Keywords/Negatives anlegen, Sponsored Brands). Jede
# Schreibaktion zeigt erst den Ist-Stand, fragt nach (ausser --ja) und schreibt dann.

def _ja(args, frage):
    if getattr(args, "ja", False):
        return
    if input(frage + " (ja/nein): ").strip().lower() != "ja":
        sys.exit("Abgebrochen. Nichts geschrieben.")


def _eine_kampagne(tenant, auswahl):
    ids, namen, _ = kampagnen_ids(tenant, [auswahl], ["ENABLED", "PAUSED", "ARCHIVED"])
    if len(ids) != 1:
        sys.exit(f"'{auswahl}' trifft {len(ids)} Kampagnen: {', '.join(namen)}. Bitte eindeutiger (oder die ID).")
    return ids[0], namen[0]


def cmd_platzierung(args):
    tenant, name = firma_id(args.firma)
    ids, namen, _ = kampagnen_ids(tenant, args.kampagne, ["ENABLED", "PAUSED"])
    d = ruf({"action": "platzierung", "company_id": tenant, "kampagnen": ids})
    print(f"Firma: {name}")
    rows = []
    for k in d["kampagnen"]:
        for pz in k["platzierungen"] or [{"placement": "(keine)", "percentage": ""}]:
            rows.append({"kampagne": k["name"], "state": k["state"], "strategie": k["strategie"], "platzierung": pz["placement"], "prozent": pz["percentage"], "campaignId": k["campaignId"]})
    tabelle(rows, ["kampagne", "state", "strategie", "platzierung", "prozent", "campaignId"])


def cmd_platzierung_setzen(args):
    tenant, name = firma_id(args.firma)
    cid, kname = _eine_kampagne(tenant, args.kampagne)
    d = ruf({"action": "platzierung", "company_id": tenant, "kampagnen": [cid]})["kampagnen"][0]
    ist = next((p["percentage"] for p in d["platzierungen"] if p["placement"] == args.placement), None)
    print(f"Firma: {name}   Kampagne: {kname} ({cid})")
    print(f"{args.placement}: aktuell {ist if ist is not None else 'nicht gesetzt'} % -> neu {args.prozent} %")
    if ist == args.prozent:
        sys.exit("Schon so. Nichts zu tun.")
    _ja(args, "Modifier bei Amazon setzen?")
    r = ruf({"action": "platzierung_setzen", "company_id": tenant, "campaignId": cid, "placement": args.placement,
             "prozent": args.prozent, "bestaetigung": True, "grund": args.grund})
    print(f"Ergebnis: {r['ergebnis']}   vorher: {r['vorher']}   nachher: {r['nachher']}")
    if r.get("detail"):
        print("Detail:", json.dumps(r["detail"], ensure_ascii=False)[:600])


def cmd_negatives(args):
    tenant, name = firma_id(args.firma)
    ids, namen, kname = kampagnen_ids(tenant, args.kampagne, ["ENABLED", "PAUSED"])
    d = ruf({"action": "negatives", "company_id": tenant, "kampagnen": ids})
    for n in d["negatives"]:
        n["kampagne"] = kname.get(n["campaignId"], n["campaignId"])
    print(f"Firma: {name}   Kampagnen: {', '.join(namen)}   Negatives: {len(d['negatives'])}")
    tabelle(sorted(d["negatives"], key=lambda n: (n["kampagne"], n["ebene"], n["text"])), ["kampagne", "ebene", "text", "matchType", "state", "keywordId"])


def cmd_keyword_anlegen(args):
    tenant, name = firma_id(args.firma)
    cid, kname = _eine_kampagne(tenant, args.kampagne)
    print(f"Firma: {name}   Kampagne: {kname} ({cid})")
    print(f"Neues Keyword: '{args.text}'  {args.match}  Gebot {args.gebot}")
    _ja(args, "Keyword bei Amazon anlegen?")
    r = ruf({"action": "keyword_anlegen", "company_id": tenant, "campaignId": cid, "adGroupId": args.adgroup,
             "keywordText": args.text, "matchType": args.match, "bid": args.gebot, "bestaetigung": True, "grund": args.grund})
    print(f"Ergebnis: {r['ergebnis']}   keywordId: {r.get('keywordId')}   " + (f"Detail: {json.dumps(r.get('detail') or r.get('keyword'), ensure_ascii=False)[:600]}" if r['ergebnis'] != 'ok' else ""))


def cmd_negative_anlegen(args):
    tenant, name = firma_id(args.firma)
    cid, kname = _eine_kampagne(tenant, args.kampagne)
    print(f"Firma: {name}   Kampagne: {kname} ({cid})")
    print(f"Neues Negative (Anzeigengruppe): '{args.text}'  {args.match}")
    _ja(args, "Negative bei Amazon anlegen?")
    r = ruf({"action": "negative_anlegen", "company_id": tenant, "campaignId": cid, "adGroupId": args.adgroup,
             "keywordText": args.text, "matchType": args.match, "bestaetigung": True, "grund": args.grund})
    print(f"Ergebnis: {r['ergebnis']}   keywordId: {r.get('keywordId')}   " + (f"Detail: {json.dumps(r.get('detail') or r.get('keyword'), ensure_ascii=False)[:600]}" if r['ergebnis'] != 'ok' else ""))


def cmd_sb_kampagnen(args):
    tenant, name = firma_id(args.firma)
    d = ruf({"action": "sb_kampagnen", "company_id": tenant, "kampagnen": args.kampagne or [], "status": args.status.split(",")})
    print(f"Firma: {name}   SB-Kampagnen: {len(d['kampagnen'])}")
    tabelle(d["kampagnen"], ["campaignId", "name", "state", "format", "budget", "budgetType"])


def cmd_sb_zustand(args):
    tenant, name = firma_id(args.firma)
    d = ruf({"action": "sb_kampagnen", "company_id": tenant, "kampagnen": [args.kampagne]})["kampagnen"]
    if not d:
        sys.exit("SB-Kampagne nicht gefunden.")
    k = d[0]
    print(f"Firma: {name}   SB-Kampagne: {k['name']} ({k['campaignId']})   Zustand: {k['state']} -> {args.state}")
    if k["state"] == args.state:
        sys.exit("Schon so. Nichts zu tun.")
    _ja(args, "Zustand bei Amazon setzen?")
    r = ruf({"action": "sb_kampagne_zustand", "company_id": tenant, "campaignId": args.kampagne, "state": args.state, "bestaetigung": True, "grund": args.grund})
    print(f"Ergebnis: {r['ergebnis']}   {r.get('vorher')} -> {r.get('nachher')}" + (f"   Detail: {json.dumps(r.get('detail'), ensure_ascii=False)[:600]}" if r.get('detail') else ""))


def cmd_sb_negatives(args):
    tenant, name = firma_id(args.firma)
    d = ruf({"action": "sb_negatives", "company_id": tenant, "campaignId": args.kampagne})
    print(f"Firma: {name}   SB-Kampagne {args.kampagne}   Negatives: {len(d['negatives'])}")
    tabelle(sorted(d["negatives"], key=lambda n: n["text"]), ["text", "matchType", "state", "keywordId"])


def cmd_sb_negatives_anlegen(args):
    tenant, name = firma_id(args.firma)
    vorhanden = ruf({"action": "sb_negatives", "company_id": tenant, "campaignId": args.kampagne})["negatives"]
    alt = {(n["text"].lower(), n["matchType"]) for n in vorhanden}
    neu = [t for t in args.text if (t.lower(), args.match) not in alt]
    dup = [t for t in args.text if (t.lower(), args.match) in alt]
    print(f"Firma: {name}   SB-Kampagne {args.kampagne}   bestehende Negatives: {len(vorhanden)}")
    print(f"Neu anzulegen ({args.match}): {neu}")
    if dup:
        print(f"Schon vorhanden, werden uebersprungen: {dup}")
    if not neu:
        sys.exit("Nichts anzulegen.")
    _ja(args, "Negatives bei Amazon anlegen?")
    r = ruf({"action": "sb_negatives_anlegen", "company_id": tenant, "campaignId": args.kampagne,
             "keywords": [{"text": t, "matchType": args.match} for t in neu], "bestaetigung": True, "grund": args.grund})
    print(f"Angelegt: {r['angelegt']}   uebersprungen: {r['uebersprungen']}   Fehler: {r['fehler']}")
    tabelle(r["ergebnisse"], ["text", "matchType", "ergebnis", "keywordId", "detail"])


# ----------------------------------------------------------------- main

def main():
    p = argparse.ArgumentParser(description="Gebote fuer Sponsored Products lesen/setzen (nur Coach, nur lokal).")
    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("login", help="einmalig anmelden (Passwort wird abgefragt)")
    s.add_argument("--email")
    s.set_defaults(fn=login)

    s = sub.add_parser("firmen", help="Firmen mit Ads-Verbindung")
    s.set_defaults(fn=cmd_firmen)

    s = sub.add_parser("kampagnen", help="Kampagnen einer Firma")
    s.add_argument("--firma", required=True)
    s.add_argument("--status", default="ENABLED,PAUSED", help="ENABLED,PAUSED,ARCHIVED")
    s.set_defaults(fn=cmd_kampagnen)

    def auswahl(s):
        s.add_argument("--firma", required=True)
        s.add_argument("--kampagne", action="append", required=True, help="campaignId oder Namensteil, mehrfach moeglich")
        s.add_argument("--kampagnen-status", default="ENABLED,PAUSED", help="welche Kampagnen beim Namens-Match zaehlen")
        s.add_argument("--status", default="ENABLED", help="Status der Keywords/Targets: ENABLED,PAUSED")
        s.add_argument("--nur", choices=["keyword", "target"], default=None, help="nur Keywords oder nur Product-Targets")
        s.add_argument("--treffer", action="append", default=None, help="nur diese Zeilen: Keyword-/Target-ID oder Textteil, mehrfach moeglich")

    s = sub.add_parser("gebote", help="aktuelle Gebote ansehen")
    auswahl(s)
    s.set_defaults(fn=cmd_gebote)

    s = sub.add_parser("vorschau", help="Aenderung berechnen, nichts schreiben")
    auswahl(s)
    g = s.add_mutually_exclusive_group(required=True)
    g.add_argument("--prozent", type=float, help="z. B. -20 (senken) oder 10 (erhoehen)")
    g.add_argument("--faktor", type=float, help="z. B. 0.8")
    g.add_argument("--absolut", type=float, help="festes Zielgebot")
    s.add_argument("--min", type=float, help="nicht unter diesen Wert (Amazon-Minimum 0.02)")
    s.add_argument("--max", type=float, help="nicht ueber diesen Wert")
    s.set_defaults(fn=cmd_vorschau)

    s = sub.add_parser("tabelle", help="Excel-Massnahmenblatt einlesen -> Vorschau (schreibt nichts)")
    s.add_argument("--firma", required=True)
    s.add_argument("--datei", default=None, help="Pfad zur .xlsx (Standard: neueste in tools/eingang)")
    s.add_argument("--blatt", default="Gebotsänderungen", help="Blattname (Standard: Gebotsänderungen)")
    s.set_defaults(fn=cmd_tabelle)

    def schreib(s):
        s.add_argument("--firma", required=True)
        s.add_argument("--grund", default=None, help="kurze Begruendung fuers Log")
        s.add_argument("--ja", action="store_true", help="ohne Rueckfrage")

    s = sub.add_parser("platzierung", help="Platzierungs-Modifier von SP-Kampagnen ansehen")
    s.add_argument("--firma", required=True)
    s.add_argument("--kampagne", action="append", required=True)
    s.set_defaults(fn=cmd_platzierung)

    s = sub.add_parser("platzierung-setzen", help="Platzierungs-Modifier einer SP-Kampagne setzen")
    schreib(s)
    s.add_argument("--kampagne", required=True, help="campaignId oder eindeutiger Namensteil")
    s.add_argument("--placement", default="PLACEMENT_TOP", choices=["PLACEMENT_TOP", "PLACEMENT_PRODUCT_PAGE", "PLACEMENT_REST_OF_SEARCH"])
    s.add_argument("--prozent", type=int, required=True, help="0-900")
    s.set_defaults(fn=cmd_platzierung_setzen)

    s = sub.add_parser("negatives", help="SP-Negatives (Anzeigengruppe + Kampagne) ansehen")
    s.add_argument("--firma", required=True)
    s.add_argument("--kampagne", action="append", required=True)
    s.set_defaults(fn=cmd_negatives)

    s = sub.add_parser("keyword-anlegen", help="SP-Keyword anlegen")
    schreib(s)
    s.add_argument("--kampagne", required=True)
    s.add_argument("--adgroup", default=None, help="adGroupId, noetig wenn die Kampagne mehrere hat")
    s.add_argument("--text", required=True)
    s.add_argument("--match", default="EXACT", choices=["EXACT", "PHRASE", "BROAD"])
    s.add_argument("--gebot", type=float, required=True)
    s.set_defaults(fn=cmd_keyword_anlegen)

    s = sub.add_parser("negative-anlegen", help="SP-Negative (Anzeigengruppe) anlegen")
    schreib(s)
    s.add_argument("--kampagne", required=True)
    s.add_argument("--adgroup", default=None)
    s.add_argument("--text", required=True)
    s.add_argument("--match", default="NEGATIVE_EXACT", choices=["NEGATIVE_EXACT", "NEGATIVE_PHRASE"])
    s.set_defaults(fn=cmd_negative_anlegen)

    s = sub.add_parser("sb-kampagnen", help="Sponsored-Brands-Kampagnen ansehen")
    s.add_argument("--firma", required=True)
    s.add_argument("--kampagne", action="append", default=None, help="campaignId, optional")
    s.add_argument("--status", default="ENABLED,PAUSED")
    s.set_defaults(fn=cmd_sb_kampagnen)

    s = sub.add_parser("sb-zustand", help="SB-Kampagne pausieren/aktivieren")
    schreib(s)
    s.add_argument("--kampagne", required=True, help="SB campaignId")
    s.add_argument("--state", required=True, choices=["PAUSED", "ENABLED"])
    s.set_defaults(fn=cmd_sb_zustand)

    s = sub.add_parser("sb-negatives", help="Negatives einer SB-Kampagne ansehen")
    s.add_argument("--firma", required=True)
    s.add_argument("--kampagne", required=True, help="SB campaignId")
    s.set_defaults(fn=cmd_sb_negatives)

    s = sub.add_parser("sb-negatives-anlegen", help="Negatives in einer SB-Kampagne anlegen")
    schreib(s)
    s.add_argument("--kampagne", required=True, help="SB campaignId")
    s.add_argument("--text", action="append", required=True, help="Keyword-Text, mehrfach moeglich")
    s.add_argument("--match", default="negativeExact", choices=["negativeExact", "negativePhrase"])
    s.set_defaults(fn=cmd_sb_negatives_anlegen)

    s = sub.add_parser("setzen", help="die letzte Vorschau bei Amazon anwenden")
    s.add_argument("--firma", required=True)
    s.add_argument("--grund", default=None, help="kurze Begruendung fuers Log")
    s.add_argument("--ja", action="store_true", help="ohne Rueckfrage")
    s.set_defaults(fn=cmd_setzen)

    args = p.parse_args()
    args.fn(args)


if __name__ == "__main__":
    main()
